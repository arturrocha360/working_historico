from openpyxl import load_workbook
from tkinter import messagebox
from Caminhos_documentos import caminho_template_Lista_equipamentos,caminho_lista_equipamentos
def Lista_equipamentos(lista_equi_full):
    workbook = load_workbook(caminho_template_Lista_equipamentos)

    # Seleciona a planilha na qual deseja adicionar linhas
    # Ou pode selecionar uma planilha específica, por exemplo: workbook['Planilha1']
    sheet = workbook['ListaEquipamentos']

    # Especifica a linha após a qual deseja adicionar linhas
    linha_de_insercao = 11  # Insere após a quinta linha, por exemplo

    # Insere uma nova linha na posição especificada
    sheet.insert_rows(linha_de_insercao)

    nova_linha = [linha[:-3] for linha in lista_equi_full] #remove as duas ultimas duas colunas para gerar  a lista de equipamentos
    for i in range(len(nova_linha)):
        for col, valor in zip(range(1, len(nova_linha[i]) + 1), nova_linha[i]):
            sheet.cell(row=linha_de_insercao+i, column=col, value=valor)
    
    # Copia a formatação da segunda linha e aplica à primeira linha
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=linha_de_insercao, column=col)._style = sheet.cell(row=linha_de_insercao + 1, column=col)._style


    # Salva as alterações no arquivo Excel
    workbook.save(caminho_lista_equipamentos)
    messagebox.showinfo("Aviso", "Lista de equipamento Gerada!")

